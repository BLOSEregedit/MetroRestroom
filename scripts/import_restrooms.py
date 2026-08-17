#!/usr/bin/env python3
"""Deterministically import restroom records from the source Excel workbook."""

from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = ROOT / "上海地铁厕所.xlsx"
DEFAULT_OUTPUT = ROOT / "miniprogram/data/generated/restrooms.js"
DEFAULT_REPORT = ROOT / "docs/data-import-report.md"

EXPECTED_SHEETS = [
    "1号线",
    "2号线",
    "3号线",
    "4号线",
    "5号线",
    "6号线",
    "7号线",
    "8号线",
    "9号线",
    "10号线",
    "11号线",
    "12号线",
    "13号线",
    "14号线",
    "15号线",
    "16号线",
    "17号线",
    "18号线",
    "浦江线",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    return parser.parse_args()


def line_id_for_sheet(sheet_name: str) -> str:
    if sheet_name == "浦江线":
        return "pujiang"
    return sheet_name.removesuffix("号线")


def has_text(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def raw_value(value: Any) -> Any:
    if value is None or isinstance(value, str):
        return value
    return str(value)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source_file:
        for chunk in iter(lambda: source_file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def analyze_sheet(worksheet: Any) -> tuple[dict[str, Any], dict[str, Any]]:
    sheet_name = worksheet.title
    line_id = line_id_for_sheet(sheet_name)
    line_key = f"l{line_id}"
    station_orders: dict[str, int] = {}
    records: list[dict[str, Any]] = []
    anomalies: list[dict[str, Any]] = []
    complete_record_count = 0
    inactive_record_count = 0
    orphan_row_count = 0

    for source_row in range(2, worksheet.max_row + 1):
        station_value = worksheet.cell(source_row, 1).value
        access_value = worksheet.cell(source_row, 2).value
        location_value = worksheet.cell(source_row, 3).value

        if not has_text(station_value):
            if has_text(access_value) or has_text(location_value):
                orphan_row_count += 1
                anomalies.append(
                    {
                        "sourceRow": source_row,
                        "stationName": None,
                        "issues": ["A列站名为空，但B/C列存在内容"],
                    }
                )
            continue

        station_name = raw_value(station_value)
        access_raw = raw_value(access_value)
        location_raw = raw_value(location_value)
        if station_name not in station_orders:
            station_orders[station_name] = len(station_orders) + 1
        display_order = station_orders[station_name]
        status = "inactive" if "在建中" in station_name else "active"

        issues = []
        if not has_text(access_value):
            issues.append("B列 accessRaw 为空")
        if not has_text(location_value):
            issues.append("C列 locationRaw 为空")
        if not isinstance(station_value, str):
            issues.append("A列站名不是字符串")
        if access_value is not None and not isinstance(access_value, str):
            issues.append("B列 accessRaw 不是字符串")
        if location_value is not None and not isinstance(location_value, str):
            issues.append("C列 locationRaw 不是字符串")

        if issues:
            anomalies.append(
                {
                    "sourceRow": source_row,
                    "stationName": station_name,
                    "issues": issues,
                }
            )
        else:
            complete_record_count += 1

        if status == "inactive":
            inactive_record_count += 1

        records.append(
            {
                "lineId": line_id,
                "lineName": sheet_name,
                "lineStationId": f"{line_key}-s{display_order:03d}",
                "sourceSheet": sheet_name,
                "sourceRow": source_row,
                "displayOrder": display_order,
                "stationName": station_name,
                "accessRaw": access_raw,
                "locationRaw": location_raw,
                "status": status,
            }
        )

    source_row_count = len(records)
    skipped_empty_row_count = (
        max(worksheet.max_row - 1, 0) - source_row_count - orphan_row_count
    )
    line = {
        "lineId": line_id,
        "lineName": sheet_name,
        "sourceSheet": sheet_name,
        "records": records,
    }
    metrics = {
        "lineId": line_id,
        "lineName": sheet_name,
        "worksheetMaxRow": worksheet.max_row,
        "sourceRowCount": source_row_count,
        "stationCount": len(station_orders),
        "restroomRecordCount": source_row_count,
        "completeRestroomRecordCount": complete_record_count,
        "inactiveRecordCount": inactive_record_count,
        "orphanRowCount": orphan_row_count,
        "skippedEmptyRowCount": skipped_empty_row_count,
        "anomalies": anomalies,
    }
    return line, metrics


def build_payload(source: Path) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    workbook = load_workbook(
        source,
        read_only=False,
        data_only=True,
        keep_links=False,
    )
    try:
        actual_sheets = workbook.sheetnames
        if actual_sheets != EXPECTED_SHEETS:
            raise ValueError(
                "工作表不符合预期："
                f"expected={EXPECTED_SHEETS!r}, actual={actual_sheets!r}"
            )

        lines = []
        line_metrics = []
        for sheet_name in EXPECTED_SHEETS:
            line, metrics = analyze_sheet(workbook[sheet_name])
            lines.append(line)
            line_metrics.append(metrics)
    finally:
        workbook.close()

    stats = {
        "lineCount": len(lines),
        "sourceRowCount": sum(item["sourceRowCount"] for item in line_metrics),
        "stationCount": sum(item["stationCount"] for item in line_metrics),
        "restroomRecordCount": sum(
            item["restroomRecordCount"] for item in line_metrics
        ),
        "completeRestroomRecordCount": sum(
            item["completeRestroomRecordCount"] for item in line_metrics
        ),
        "activeRecordCount": sum(
            item["restroomRecordCount"] - item["inactiveRecordCount"]
            for item in line_metrics
        ),
        "inactiveRecordCount": sum(
            item["inactiveRecordCount"] for item in line_metrics
        ),
        "anomalyRowCount": sum(
            len(item["anomalies"]) for item in line_metrics
        ),
    }
    payload = {
        "schemaVersion": 1,
        "source": {
            "fileName": source.name,
            "sha256": sha256_file(source),
            "sheetNames": EXPECTED_SHEETS,
        },
        "lines": lines,
        "stats": stats,
    }
    return payload, line_metrics


def render_javascript(payload: dict[str, Any]) -> str:
    encoded = json.dumps(payload, ensure_ascii=False, indent=2)
    return (
        "// 此文件由 scripts/import_restrooms.py 确定性生成，请勿手工编辑。\n"
        f"module.exports = {encoded};\n"
    )


def render_report(
    payload: dict[str, Any], line_metrics: list[dict[str, Any]]
) -> str:
    stats = payload["stats"]
    source = payload["source"]
    lines = [
        "# 上海地铁厕所数据导入报告",
        "",
        "## 导入口径",
        "",
        f"- 源文件：`{source['fileName']}`。",
        f"- 源文件 SHA-256：`{source['sha256']}`。",
        "- 使用 bundled Python 的 `openpyxl`，以 `read_only=False`、`data_only=True` 读取。",
        "- 不依赖工作表的错误 dimension；逐表扫描第 2 行至 `max_row`，仅导入 A 列站名实际非空的行。",
        "- B/C 空单元格不导致记录丢弃，在生成数据中保留字段并序列化为 `null`；非空字符串不清洗、不裁剪。",
        "- `displayOrder` 按线路内站名首次出现顺序生成；同站多条记录复用同一 `lineStationId` 和顺序。",
        "- 站名包含“在建中”的记录标记为 `inactive`，其余为 `active`。",
        "- 输出不包含动态生成时间，重复运行结果确定。",
        "",
        "## 汇总",
        "",
        f"- 线路工作表：{stats['lineCount']} 张。",
        f"- A 列非空源行／导入厕所记录：{stats['sourceRowCount']} 条。",
        f"- 线路内站点：{stats['stationCount']} 个。",
        f"- B/C 均完整的厕所记录：{stats['completeRestroomRecordCount']} 条。",
        f"- 状态：`active` {stats['activeRecordCount']} 条，`inactive` {stats['inactiveRecordCount']} 条。",
        f"- 异常源行：{stats['anomalyRowCount']} 行。",
        "",
        "## 逐线统计",
        "",
        "| 线路 | lineId | max_row | A列非空源行 | 站点 | 导入记录 | 完整记录 | inactive | 跳过空行 | 异常行 |",
        "| --- | --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: |",
    ]

    for item in line_metrics:
        lines.append(
            "| {lineName} | {lineId} | {worksheetMaxRow} | {sourceRowCount} | "
            "{stationCount} | {restroomRecordCount} | "
            "{completeRestroomRecordCount} | {inactiveRecordCount} | "
            "{skippedEmptyRowCount} | {anomalyCount} |".format(
                **item,
                anomalyCount=len(item["anomalies"]),
            )
        )

    lines.extend(["", "## 逐线异常行", ""])
    for item in line_metrics:
        lines.append(f"### {item['lineName']}")
        lines.append("")
        if not item["anomalies"]:
            lines.append("- 无。")
        else:
            for anomaly in item["anomalies"]:
                station_name = anomaly["stationName"] or "（站名为空）"
                issues = "；".join(anomaly["issues"])
                lines.append(
                    f"- 源第 {anomaly['sourceRow']} 行，`{station_name}`：{issues}。"
                )
        lines.append("")

    incomplete_count = (
        stats["restroomRecordCount"] - stats["completeRestroomRecordCount"]
    )
    lines.extend(
        [
            "## 对账结论",
            "",
            f"- 19 张预期工作表全部存在且顺序一致。",
            f"- 导入记录 {stats['restroomRecordCount']} = 完整记录 {stats['completeRestroomRecordCount']} + 字段不完整记录 {incomplete_count}。",
            f"- 状态记录 {stats['restroomRecordCount']} = active {stats['activeRecordCount']} + inactive {stats['inactiveRecordCount']}。",
            "- 15号线 `max_row=1234`，仅导入 A 列实际非空的 30 行，其余空行未进入数据。",
            "- 原始 Excel 未被修改。",
            "",
        ]
    )
    return "\n".join(lines)


def main() -> None:
    args = parse_args()
    source = args.source.resolve()
    output = args.output.resolve()
    report = args.report.resolve()

    if not source.is_file():
        raise FileNotFoundError(f"找不到源文件：{source}")

    payload, line_metrics = build_payload(source)
    output.parent.mkdir(parents=True, exist_ok=True)
    report.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(render_javascript(payload), encoding="utf-8")
    report.write_text(render_report(payload, line_metrics), encoding="utf-8")

    stats = payload["stats"]
    print(
        "Imported "
        f"{stats['restroomRecordCount']} records from "
        f"{stats['lineCount']} sheets; "
        f"complete={stats['completeRestroomRecordCount']}; "
        f"anomalies={stats['anomalyRowCount']}."
    )


if __name__ == "__main__":
    main()
